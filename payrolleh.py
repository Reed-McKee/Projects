from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font

wb = load_workbook('payrollvalues.xlsx')
ws = wb.active

start_row = 5
names = []
hour_row = 22
hours =[]

payscale = 20.00

while ws.cell(row=start_row, column=3).value:
    name = ws.cell(row=start_row, column=3).value
    names.append(name)
    start_row += 22 

while ws.cell(row=hour_row, column=11).value is not None:
    hour = ws.cell(row=hour_row, column=11).value
    hours.append(hour)
    hour_row +=22

ws.delete_rows(1, ws.max_row)

for index, name in enumerate(names, start=1):
    realname = name[5:]
    ws.cell(row=index, column=1, value=realname)

for hour_index, hour in enumerate(hours, start=1):
    ws.cell(row=hour_index, column=2, value=hour)  
    regular_hours = min(hour, 80)
    overtime_hours = max(hour - 80, 0)
    ws.cell(row=hour_index, column=3, value=regular_hours)
    ws.cell(row=hour_index, column=4, value=overtime_hours) 
    regular_pay = regular_hours * payscale 
    overtime_pay = overtime_hours * (payscale * 1.5)
    total_pay = regular_pay + overtime_pay 
    ws.cell(row=hour_index, column=5, value=regular_pay)
    ws.cell(row=hour_index, column=6, value=overtime_pay)
    ws.cell(row=hour_index, column=7, value=total_pay)
    rounded_total_pay = round(total_pay)  # Round to nearest dollar
    ws.cell(row=hour_index, column=8, value=f"${rounded_total_pay}")

ws.insert_rows(1)  # Add a row at the top for titles
ws.cell(row=1, column=1, value="Name")
ws.cell(row=1, column=2, value="Hours Worked")
ws.cell(row=1, column=3, value="Regular Hours")
ws.cell(row=1, column=4, value="Overtime Hours")
ws.cell(row=1, column=5, value="Regular Pay")
ws.cell(row=1, column=6, value="Overtime Pay")
ws.cell(row=1, column=7, value="Total Pay")
ws.cell(row=1, column=8, value="Adjusted Pay")

ws.column_dimensions['A'].width = max(len(name) for name in names) + 2
ws.column_dimensions['B'].width = len("Hours Worked") + 2
ws.column_dimensions['C'].width = len("Regular Hours") + 2
ws.column_dimensions['D'].width = len("Overtime Hours") + 2
ws.column_dimensions['E'].width = len("Regular Pay") + 2
ws.column_dimensions['F'].width = len("Overtime Pay") + 2
ws.column_dimensions['G'].width = len("Total Pay") + 2
ws.column_dimensions['H'].width = len("Adjusted Pay") + 2

title_alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].alignment = title_alignment
ws['B1'].alignment = title_alignment
ws['C1'].alignment = title_alignment
ws['D1'].alignment = title_alignment
ws['E1'].alignment = title_alignment
ws['F1'].alignment = title_alignment
ws['G1'].alignment = title_alignment
ws['H1'].alignment = title_alignment



wb.save('test43.xlsx')